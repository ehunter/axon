"""Export service for generating sample selection reports."""

import csv
import json
from dataclasses import dataclass, field
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Any

from axon.agent.tools import SampleSelection, SelectedSample
from axon.export.formats import ExportFormat


@dataclass
class ExportMetadata:
    """Metadata about the export."""
    researcher_name: str | None = None
    research_purpose: str | None = None
    tissue_use: str | None = None  # e.g., "RNA-seq", "Histology"
    selection_criteria: dict[str, Any] = field(default_factory=dict)
    notes: str | None = None
    exported_at: datetime = field(default_factory=datetime.now)


@dataclass 
class ExportResult:
    """Result of an export operation."""
    format: ExportFormat
    filename: str
    content: str | bytes
    sample_count: int
    metadata: ExportMetadata


class ExportService:
    """Service for exporting sample selections to various formats."""
    
    def __init__(self, selection: SampleSelection, metadata: ExportMetadata | None = None):
        """Initialize the export service.
        
        Args:
            selection: The current sample selection
            metadata: Optional metadata about the export
        """
        self.selection = selection
        self.metadata = metadata or ExportMetadata()
    
    def export(self, format: ExportFormat, output_path: Path | None = None) -> ExportResult:
        """Export the selection to the specified format.
        
        Args:
            format: The export format
            output_path: Optional path to write the file
            
        Returns:
            ExportResult with the generated content
        """
        if format == ExportFormat.CSV:
            result = self._export_csv()
        elif format == ExportFormat.EXCEL:
            result = self._export_excel()
        elif format == ExportFormat.JSON:
            result = self._export_json()
        elif format == ExportFormat.TEXT:
            result = self._export_text()
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        if output_path:
            self._write_to_file(result, output_path)
        
        return result
    
    def _export_csv(self) -> ExportResult:
        """Export to CSV format."""
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Group",
            "Sample ID",
            "Repository",
            "Diagnosis",
            "Age",
            "Sex",
            "RIN",
            "PMI (hours)",
            "Brain Region",
            "Braak Stage",
            "Co-pathologies",
        ])
        
        # Cases
        for sample in self.selection.cases:
            writer.writerow(self._sample_to_row(sample, "Case"))
        
        # Controls
        for sample in self.selection.controls:
            writer.writerow(self._sample_to_row(sample, "Control"))
        
        content = output.getvalue()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return ExportResult(
            format=ExportFormat.CSV,
            filename=f"brain_samples_{timestamp}.csv",
            content=content,
            sample_count=len(self.selection.cases) + len(self.selection.controls),
            metadata=self.metadata,
        )
    
    def _export_excel(self) -> ExportResult:
        """Export to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        self._build_summary_sheet(summary_sheet, Font, PatternFill, Alignment)
        
        # Samples sheet
        samples_sheet = wb.create_sheet("Samples")
        self._build_samples_sheet(samples_sheet, Font, PatternFill, Alignment, Border, Side, get_column_letter)
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        content = output.getvalue()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return ExportResult(
            format=ExportFormat.EXCEL,
            filename=f"brain_samples_{timestamp}.xlsx",
            content=content,
            sample_count=len(self.selection.cases) + len(self.selection.controls),
            metadata=self.metadata,
        )
    
    def _build_summary_sheet(self, sheet, Font, PatternFill, Alignment):
        """Build the summary sheet for Excel export."""
        # Title
        sheet["A1"] = "Brain Sample Selection Summary"
        sheet["A1"].font = Font(bold=True, size=16)
        sheet.merge_cells("A1:D1")
        
        # Metadata
        row = 3
        sheet[f"A{row}"] = "Export Date:"
        sheet[f"B{row}"] = self.metadata.exported_at.strftime("%Y-%m-%d %H:%M")
        
        row += 1
        sheet[f"A{row}"] = "Total Samples:"
        sheet[f"B{row}"] = len(self.selection.cases) + len(self.selection.controls)
        
        row += 1
        sheet[f"A{row}"] = "Cases:"
        sheet[f"B{row}"] = len(self.selection.cases)
        
        row += 1
        sheet[f"A{row}"] = "Controls:"
        sheet[f"B{row}"] = len(self.selection.controls)
        
        if self.metadata.research_purpose:
            row += 2
            sheet[f"A{row}"] = "Research Purpose:"
            sheet[f"B{row}"] = self.metadata.research_purpose
        
        if self.metadata.tissue_use:
            row += 1
            sheet[f"A{row}"] = "Tissue Use:"
            sheet[f"B{row}"] = self.metadata.tissue_use
        
        if self.metadata.selection_criteria:
            row += 2
            sheet[f"A{row}"] = "Selection Criteria:"
            sheet[f"A{row}"].font = Font(bold=True)
            
            for key, value in self.metadata.selection_criteria.items():
                row += 1
                sheet[f"A{row}"] = f"  {key}:"
                sheet[f"B{row}"] = str(value)
        
        if self.metadata.notes:
            row += 2
            sheet[f"A{row}"] = "Notes:"
            sheet[f"B{row}"] = self.metadata.notes
        
        # Column widths
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 40
    
    def _build_samples_sheet(self, sheet, Font, PatternFill, Alignment, Border, Side, get_column_letter):
        """Build the samples sheet for Excel export."""
        headers = [
            "Group",
            "Sample ID",
            "Repository",
            "Diagnosis",
            "Age",
            "Sex",
            "RIN",
            "PMI (hours)",
            "Brain Region",
            "Braak Stage",
            "Co-pathologies",
        ]
        
        # Header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        row = 2
        case_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        control_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
        
        for sample in self.selection.cases:
            data = self._sample_to_row(sample, "Case")
            for col, value in enumerate(data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.fill = case_fill
            row += 1
        
        for sample in self.selection.controls:
            data = self._sample_to_row(sample, "Control")
            for col, value in enumerate(data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.fill = control_fill
            row += 1
        
        # Column widths
        col_widths = [10, 15, 20, 35, 8, 8, 8, 12, 25, 15, 30]
        for col, width in enumerate(col_widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header
        sheet.freeze_panes = "A2"
    
    def _export_json(self) -> ExportResult:
        """Export to JSON format."""
        data = {
            "metadata": {
                "exported_at": self.metadata.exported_at.isoformat(),
                "researcher_name": self.metadata.researcher_name,
                "research_purpose": self.metadata.research_purpose,
                "tissue_use": self.metadata.tissue_use,
                "selection_criteria": self.metadata.selection_criteria,
                "notes": self.metadata.notes,
            },
            "summary": {
                "total_samples": len(self.selection.cases) + len(self.selection.controls),
                "case_count": len(self.selection.cases),
                "control_count": len(self.selection.controls),
            },
            "cases": [self._sample_to_dict(s) for s in self.selection.cases],
            "controls": [self._sample_to_dict(s) for s in self.selection.controls],
        }
        
        content = json.dumps(data, indent=2)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return ExportResult(
            format=ExportFormat.JSON,
            filename=f"brain_samples_{timestamp}.json",
            content=content,
            sample_count=len(self.selection.cases) + len(self.selection.controls),
            metadata=self.metadata,
        )
    
    def _export_text(self) -> ExportResult:
        """Export to human-readable text format (for email/printing)."""
        lines = []
        
        # Header
        lines.append("=" * 70)
        lines.append("BRAIN SAMPLE SELECTION SUMMARY")
        lines.append("=" * 70)
        lines.append("")
        
        # Metadata
        lines.append(f"Export Date: {self.metadata.exported_at.strftime('%Y-%m-%d %H:%M')}")
        lines.append(f"Total Samples: {len(self.selection.cases) + len(self.selection.controls)}")
        lines.append(f"  - Cases: {len(self.selection.cases)}")
        lines.append(f"  - Controls: {len(self.selection.controls)}")
        lines.append("")
        
        if self.metadata.research_purpose:
            lines.append(f"Research Purpose: {self.metadata.research_purpose}")
        
        if self.metadata.tissue_use:
            lines.append(f"Tissue Use: {self.metadata.tissue_use}")
        
        if self.metadata.selection_criteria:
            lines.append("")
            lines.append("Selection Criteria:")
            for key, value in self.metadata.selection_criteria.items():
                lines.append(f"  - {key}: {value}")
        
        if self.metadata.notes:
            lines.append("")
            lines.append(f"Notes: {self.metadata.notes}")
        
        # Cases
        lines.append("")
        lines.append("-" * 70)
        lines.append("CASE SAMPLES")
        lines.append("-" * 70)
        
        for i, sample in enumerate(self.selection.cases, 1):
            lines.extend(self._format_sample_text(i, sample))
        
        if not self.selection.cases:
            lines.append("  (No cases selected)")
        
        # Controls
        lines.append("")
        lines.append("-" * 70)
        lines.append("CONTROL SAMPLES")
        lines.append("-" * 70)
        
        for i, sample in enumerate(self.selection.controls, 1):
            lines.extend(self._format_sample_text(i, sample))
        
        if not self.selection.controls:
            lines.append("  (No controls selected)")
        
        # Footer
        lines.append("")
        lines.append("=" * 70)
        lines.append("END OF REPORT")
        lines.append("=" * 70)
        
        content = "\n".join(lines)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return ExportResult(
            format=ExportFormat.TEXT,
            filename=f"brain_samples_{timestamp}.txt",
            content=content,
            sample_count=len(self.selection.cases) + len(self.selection.controls),
            metadata=self.metadata,
        )
    
    def _sample_to_row(self, sample: SelectedSample, group: str) -> list:
        """Convert a sample to a CSV/Excel row."""
        return [
            group,
            sample.external_id,
            sample.repository,
            sample.diagnosis,
            sample.age if sample.age else "N/A",
            sample.sex if sample.sex else "N/A",
            f"{sample.rin:.1f}" if sample.rin else "N/A",
            f"{sample.pmi:.1f}" if sample.pmi else "N/A",
            sample.brain_region if sample.brain_region else "N/A",
            sample.braak_stage if sample.braak_stage else "N/A",
            sample.copathologies if sample.copathologies else "None recorded",
        ]
    
    def _sample_to_dict(self, sample: SelectedSample) -> dict:
        """Convert a sample to a dictionary."""
        return {
            "external_id": sample.external_id,
            "repository": sample.repository,
            "diagnosis": sample.diagnosis,
            "age": sample.age,
            "sex": sample.sex,
            "rin": sample.rin,
            "pmi": sample.pmi,
            "brain_region": sample.brain_region,
            "braak_stage": sample.braak_stage,
            "copathologies": sample.copathologies,
        }
    
    def _format_sample_text(self, index: int, sample: SelectedSample) -> list[str]:
        """Format a sample for text output."""
        lines = []
        lines.append("")
        lines.append(f"  {index}. {sample.external_id} ({sample.repository})")
        lines.append(f"     Diagnosis: {sample.diagnosis}")
        lines.append(f"     Demographics: {sample.age or 'N/A'}yo {sample.sex or 'N/A'}")
        
        quality_parts = []
        if sample.rin:
            quality_parts.append(f"RIN: {sample.rin:.1f}")
        if sample.pmi:
            quality_parts.append(f"PMI: {sample.pmi:.1f}h")
        if quality_parts:
            lines.append(f"     Quality: {', '.join(quality_parts)}")
        
        if sample.brain_region:
            lines.append(f"     Brain Region: {sample.brain_region}")
        
        if sample.braak_stage:
            lines.append(f"     Braak Stage: {sample.braak_stage}")
        
        if sample.copathologies:
            lines.append(f"     Co-pathologies: {sample.copathologies}")
        
        return lines
    
    def _write_to_file(self, result: ExportResult, output_path: Path) -> None:
        """Write the export result to a file."""
        if isinstance(result.content, bytes):
            output_path.write_bytes(result.content)
        else:
            output_path.write_text(result.content)
    
    def generate_admin_email(self) -> str:
        """Generate an email summary for brain bank administrators.
        
        Returns:
            Formatted email text
        """
        lines = []
        
        lines.append("Subject: Brain Sample Request")
        lines.append("")
        lines.append("Dear Brain Bank Administrator,")
        lines.append("")
        lines.append("A researcher has submitted a tissue sample request through the Axon system.")
        lines.append("")
        
        # Request details
        lines.append("REQUEST DETAILS")
        lines.append("-" * 40)
        lines.append(f"Date: {self.metadata.exported_at.strftime('%Y-%m-%d %H:%M')}")
        
        if self.metadata.researcher_name:
            lines.append(f"Researcher: {self.metadata.researcher_name}")
        
        if self.metadata.research_purpose:
            lines.append(f"Research Purpose: {self.metadata.research_purpose}")
        
        if self.metadata.tissue_use:
            lines.append(f"Intended Use: {self.metadata.tissue_use}")
        
        lines.append("")
        lines.append(f"Total Samples Requested: {len(self.selection.cases) + len(self.selection.controls)}")
        lines.append(f"  - Cases: {len(self.selection.cases)}")
        lines.append(f"  - Controls: {len(self.selection.controls)}")
        lines.append("")
        
        # Selection criteria
        if self.metadata.selection_criteria:
            lines.append("SELECTION CRITERIA")
            lines.append("-" * 40)
            for key, value in self.metadata.selection_criteria.items():
                lines.append(f"  {key}: {value}")
            lines.append("")
        
        # Sample list by repository
        repos = {}
        for sample in self.selection.cases:
            repo = sample.repository or "Unknown"
            if repo not in repos:
                repos[repo] = {"cases": [], "controls": []}
            repos[repo]["cases"].append(sample)
        
        for sample in self.selection.controls:
            repo = sample.repository or "Unknown"
            if repo not in repos:
                repos[repo] = {"cases": [], "controls": []}
            repos[repo]["controls"].append(sample)
        
        lines.append("SAMPLES BY REPOSITORY")
        lines.append("-" * 40)
        
        for repo, samples in sorted(repos.items()):
            lines.append("")
            lines.append(f"  {repo}")
            
            if samples["cases"]:
                lines.append(f"    Cases ({len(samples['cases'])}):")
                for s in samples["cases"]:
                    lines.append(f"      - {s.external_id}: {s.diagnosis}, {s.age}yo {s.sex}")
            
            if samples["controls"]:
                lines.append(f"    Controls ({len(samples['controls'])}):")
                for s in samples["controls"]:
                    lines.append(f"      - {s.external_id}: {s.diagnosis}, {s.age}yo {s.sex}")
        
        lines.append("")
        lines.append("-" * 40)
        lines.append("")
        lines.append("Please review this request and contact the researcher to discuss availability")
        lines.append("and next steps.")
        lines.append("")
        lines.append("Best regards,")
        lines.append("Axon Brain Bank Discovery System")
        
        return "\n".join(lines)

